"""
Simple Excel reader for schedule rows.

Expects an XLSX with a header row. Common header names:
  email, date, time, timezone, latitude, longitude

Returns a list of dicts (one per non-empty row) with normalized header keys.
"""
from openpyxl import load_workbook
from typing import List, Dict, Any
import os

def _normalize_header(h: str) -> str:
    if h is None:
        return ""
    return str(h).strip().lower().replace(" ", "_")

def read_schedules(path: str) -> List[Dict[str, Any]]:
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel file not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_normalize_header(h) for h in rows[0]]
    schedules = []
    for row in rows[1:]:
        # skip entirely empty rows
        if all(cell is None for cell in row):
            continue
        # zip headers and row; if header is empty skip that column
        item = {}
        for h, val in zip(headers, row):
            if not h:
                continue
            item[h] = val
        schedules.append(item)
    return schedules

# quick manual test when run directly
if __name__ == "__main__":
    import sys
    p = sys.argv[1] if len(sys.argv) > 1 else "data/sample_schedules.xlsx"
    print(read_schedules(p))
